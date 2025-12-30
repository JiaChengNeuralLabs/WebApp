"""
Comando para importar alumnos, pagos y facturas trimestrales desde archivos Excel Trimestre-X.xlsx

Uso:
    python manage.py import_trimestre C:\\path\\to\\Trimestre-1.xlsx
    python manage.py import_trimestre C:\\path\\to\\Trimestre-1.xlsx --dry-run

El comando:
1. Crea alumnos nuevos si no existen (busca por DNI)
2. Actualiza direccion de alumnos existentes si esta vacia
3. Registra pagos con el TOTAL de cada fila del Excel
4. Crea facturas trimestrales (TaxInvoice) con los datos de BASE, IVA, TASAS

Columnas esperadas del Excel:
    A: CURSO, B: N FACTURA, C: FECHA, D: NOMBRE Y APELLIDOS, E: DNI,
    F: BASE IMPONIBLE, G: IVA, H: TASAS, I: TOTAL,
    J: DIRECCION, K: CP, L: MUNICIPIO, M: PROVINCIA
"""

from django.core.management.base import BaseCommand, CommandError
from students.models import Student, Payment, LicenseType, TaxInvoice
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime
import re


class Command(BaseCommand):
    help = 'Importa alumnos, pagos y facturas trimestrales desde un archivo Excel (Trimestre-X.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta al archivo Excel')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Simular importacion sin guardar en la base de datos'
        )

    def normalize_dni(self, dni):
        """Normaliza el DNI para comparacion."""
        if not dni:
            return ''
        return str(dni).upper().replace(' ', '').replace('-', '').strip()

    def parse_name(self, full_name):
        """Separa nombre completo en nombre y apellidos."""
        if not full_name:
            return '', ''

        full_name = str(full_name).strip()
        parts = full_name.split()

        if len(parts) == 0:
            return '', ''
        elif len(parts) == 1:
            return parts[0], ''
        elif len(parts) == 2:
            # Asumimos: APELLIDO NOMBRE o NOMBRE APELLIDO
            # En España es más común APELLIDOS NOMBRE, así que:
            return parts[1], parts[0]  # nombre, apellido
        elif len(parts) == 3:
            # APELLIDO1 APELLIDO2 NOMBRE
            return parts[2], f"{parts[0]} {parts[1]}"
        else:
            # APELLIDO1 APELLIDO2 NOMBRE1 NOMBRE2...
            return ' '.join(parts[2:]), f"{parts[0]} {parts[1]}"

    def parse_amount(self, value):
        """Convierte valor a Decimal."""
        if value is None:
            return Decimal('0.00')

        try:
            # Si es número, convertir directamente
            if isinstance(value, (int, float)):
                return Decimal(str(value)).quantize(Decimal('0.01'))

            # Si es string, limpiar y convertir
            value_str = str(value).strip()
            # Remover símbolo de euro y espacios
            value_str = value_str.replace('€', '').replace(' ', '')
            # Manejar formato español (1.234,56) vs inglés (1,234.56)
            if ',' in value_str and '.' in value_str:
                # Formato español: 1.234,56
                value_str = value_str.replace('.', '').replace(',', '.')
            elif ',' in value_str:
                # Solo coma: puede ser 1234,56 (español) o 1,234 (inglés miles)
                # Asumimos español si hay 2 decimales después de la coma
                parts = value_str.split(',')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    value_str = value_str.replace(',', '.')
                else:
                    value_str = value_str.replace(',', '')

            return Decimal(value_str).quantize(Decimal('0.01'))
        except (InvalidOperation, ValueError):
            return Decimal('0.00')

    def parse_date(self, value):
        """Convierte valor a fecha."""
        if value is None:
            return None

        # Si ya es datetime
        if isinstance(value, datetime):
            return value

        # Si es date
        if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
            return datetime(value.year, value.month, value.day)

        # Si es string, intentar parsear
        value_str = str(value).strip()

        # Formatos comunes
        formats = [
            '%d/%m/%Y',
            '%d-%m-%Y',
            '%Y-%m-%d',
            '%d/%m/%y',
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value_str, fmt)
            except ValueError:
                continue

        return None

    def get_license_type(self, curso):
        """Obtiene o crea el tipo de carnet."""
        if not curso:
            curso = 'B'  # Default

        curso = str(curso).upper().strip()

        # Mapeo de cursos del Excel a tipos de carnet
        mapping = {
            'AM': 'AM',
            'A1': 'A1',
            'A2': 'A2',
            'A': 'A',
            'B': 'B',
            'C': 'C',
            'C+E': 'C+E',
            'CE': 'C+E',
        }

        license_name = mapping.get(curso, curso)

        license_type, created = LicenseType.objects.get_or_create(
            name=license_name,
            defaults={'description': f'Carnet tipo {license_name}'}
        )

        return license_type

    def get_curso_code(self, curso):
        """Obtiene el código de curso para TaxInvoice."""
        if not curso:
            return 'B'

        curso = str(curso).upper().strip()

        mapping = {
            'AM': 'AM',
            'A1': 'A1',
            'A2': 'A2',
            'A': 'A',
            'B': 'B',
            'C': 'C',
            'C+E': 'C+E',
            'CE': 'C+E',
        }

        return mapping.get(curso, 'B')

    def detect_tasas(self, tasas_amount):
        """
        Detecta qué tasas están incluidas basándose en el importe total de tasas.
        Tasas DGT:
        - Tasa Básica: 94.05€
        - Tasa A (motos): 28.87€
        - Traslado: 8.67€
        - Renovación: 94.05€ (misma que básica)
        """
        if tasas_amount <= 0:
            return False, False, False, 0

        # Constantes de tasas
        TASA_BASICA = Decimal('94.05')
        TASA_A = Decimal('28.87')
        TRASLADO = Decimal('8.67')

        has_tasa_basica = False
        has_tasa_a = False
        has_traslado = False
        renovaciones = 0

        remaining = tasas_amount

        # Detectar traslado primero (es el más pequeño y específico)
        if remaining >= TRASLADO and (remaining - TRASLADO) % TASA_BASICA == 0:
            # Probablemente hay traslado
            pass  # Lo verificamos después

        # Detectar tasa A (motos)
        if remaining >= TASA_A:
            # Verificar si el resto es múltiplo de tasa básica o coincide
            test_remaining = remaining - TASA_A
            if test_remaining >= 0 and test_remaining % TASA_BASICA == 0:
                has_tasa_a = True
                remaining = test_remaining

        # Detectar traslado
        if remaining >= TRASLADO:
            test_remaining = remaining - TRASLADO
            if test_remaining >= 0 and test_remaining % TASA_BASICA == 0:
                has_traslado = True
                remaining = test_remaining

        # El resto son tasas básicas / renovaciones
        if remaining > 0:
            num_basicas = int(remaining / TASA_BASICA)
            if num_basicas > 0 and remaining == num_basicas * TASA_BASICA:
                if num_basicas == 1:
                    has_tasa_basica = True
                else:
                    has_tasa_basica = True
                    renovaciones = num_basicas - 1

        # Verificación: si no detectamos nada pero hay tasas, asumir tasa básica
        if tasas_amount > 0 and not has_tasa_basica and not has_tasa_a and not has_traslado:
            # Aproximación: si es cercano a 94.05, es tasa básica
            if abs(tasas_amount - TASA_BASICA) < Decimal('1.00'):
                has_tasa_basica = True
            elif tasas_amount > TASA_BASICA:
                has_tasa_basica = True
                renovaciones = max(0, int((tasas_amount - TASA_BASICA) / TASA_BASICA))

        return has_tasa_basica, has_tasa_a, has_traslado, renovaciones

    def get_quarter_from_date(self, date):
        """Retorna el trimestre (1-4) desde una fecha."""
        if not date:
            return 1
        month = date.month
        if 1 <= month <= 3:
            return 1
        elif 4 <= month <= 6:
            return 2
        elif 7 <= month <= 9:
            return 3
        return 4

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError('openpyxl es requerido. Instalar con: pip install openpyxl')

        excel_path = Path(options['excel_file'])
        dry_run = options['dry_run']

        if not excel_path.exists():
            raise CommandError(f'Archivo no encontrado: {excel_path}')

        self.stdout.write(f'Leyendo {excel_path.name}...')
        if dry_run:
            self.stdout.write(self.style.WARNING('MODO SIMULACION - No se guardaran cambios'))

        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active

        # Contadores
        students_created = 0
        students_updated = 0
        payments_created = 0
        payments_skipped = 0
        invoices_created = 0
        invoices_skipped = 0
        rows_skipped = 0

        # Columnas (1-indexed para openpyxl):
        # A(1): CURSO, B(2): N FACTURA, C(3): FECHA, D(4): NOMBRE Y APELLIDOS, E(5): DNI,
        # F(6): BASE IMPONIBLE, G(7): IVA, H(8): TASAS, I(9): TOTAL,
        # J(10): DIRECCION, K(11): CP, L(12): MUNICIPIO, M(13): PROVINCIA

        for row_num in range(2, ws.max_row + 1):
            curso = ws.cell(row=row_num, column=1).value
            n_factura = ws.cell(row=row_num, column=2).value
            fecha = ws.cell(row=row_num, column=3).value
            nombre_completo = ws.cell(row=row_num, column=4).value
            dni = ws.cell(row=row_num, column=5).value
            base_imponible = ws.cell(row=row_num, column=6).value
            iva = ws.cell(row=row_num, column=7).value
            tasas = ws.cell(row=row_num, column=8).value
            total = ws.cell(row=row_num, column=9).value
            direccion = ws.cell(row=row_num, column=10).value
            cp = ws.cell(row=row_num, column=11).value
            municipio = ws.cell(row=row_num, column=12).value
            provincia = ws.cell(row=row_num, column=13).value

            # Validar datos mínimos
            dni_clean = self.normalize_dni(dni)
            if not dni_clean:
                rows_skipped += 1
                continue

            # Parsear datos
            first_name, last_name = self.parse_name(nombre_completo)
            total_amount = self.parse_amount(total)
            base_amount = self.parse_amount(base_imponible)
            iva_amount = self.parse_amount(iva)
            tasas_amount = self.parse_amount(tasas)
            fecha_parsed = self.parse_date(fecha)
            n_factura_str = str(n_factura).strip() if n_factura else None

            if not first_name and not last_name:
                self.stdout.write(self.style.WARNING(
                    f'  Fila {row_num}: Sin nombre para DNI {dni_clean}, saltando...'
                ))
                rows_skipped += 1
                continue

            # Buscar o crear alumno
            student = None
            try:
                student = Student.objects.get(dni__iexact=dni_clean)

                # Actualizar direccion si esta vacia
                updated_fields = []
                if not student.street_address and direccion:
                    student.street_address = str(direccion).strip()
                    updated_fields.append('direccion')
                if not student.postal_code and cp:
                    student.postal_code = str(cp).strip()
                    updated_fields.append('CP')
                if not student.municipality and municipio:
                    student.municipality = str(municipio).strip()
                    updated_fields.append('municipio')
                if not student.province and provincia:
                    student.province = str(provincia).strip()
                    updated_fields.append('provincia')

                if updated_fields and not dry_run:
                    student.save()
                    students_updated += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'  Actualizado: {student} - {", ".join(updated_fields)}'
                    ))
                elif updated_fields and dry_run:
                    students_updated += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'  [DRY] Actualizaria: {first_name} {last_name} - {", ".join(updated_fields)}'
                    ))

            except Student.DoesNotExist:
                # Crear alumno nuevo
                license_type = self.get_license_type(curso)

                if not dry_run:
                    student = Student.objects.create(
                        first_name=first_name,
                        last_name=last_name,
                        dni=dni_clean,
                        phone='',  # No disponible en Excel
                        license_type=license_type,
                        street_address=str(direccion).strip() if direccion else '',
                        postal_code=str(cp).strip() if cp else '',
                        municipality=str(municipio).strip() if municipio else '',
                        province=str(provincia).strip() if provincia else 'VALENCIA',
                    )
                    students_created += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'  Creado alumno: {student} ({dni_clean}) - {license_type.name}'
                    ))
                else:
                    students_created += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'  [DRY] Crearia alumno: {first_name} {last_name} ({dni_clean})'
                    ))

            except Student.MultipleObjectsReturned:
                self.stdout.write(self.style.ERROR(
                    f'  Fila {row_num}: Multiples alumnos con DNI {dni_clean}'
                ))
                rows_skipped += 1
                continue

            # Registrar pago si hay importe
            payment = None
            if total_amount > 0 and not dry_run and student:
                # Verificar duplicados
                existing_payment = None

                # 1. Buscar por número de factura en las notas
                if n_factura_str:
                    existing_payment = Payment.objects.filter(
                        student=student,
                        notes__icontains=f'Factura {n_factura_str}'
                    ).first()

                # 2. Si no hay número de factura, buscar por importe+fecha
                if not existing_payment and fecha_parsed:
                    existing_payment = Payment.objects.filter(
                        student=student,
                        amount=total_amount,
                        date_paid__date=fecha_parsed.date() if hasattr(fecha_parsed, 'date') else fecha_parsed
                    ).first()

                if not existing_payment:
                    payment = Payment.objects.create(
                        student=student,
                        amount=total_amount,
                        payment_method='CARD',
                        date_paid=fecha_parsed or datetime.now(),
                        notes=f'Importado de {excel_path.name} - Factura {n_factura_str or "N/A"}'
                    )
                    payments_created += 1
                    self.stdout.write(f'    + Pago: {total_amount}€ - Factura {n_factura_str or "N/A"}')
                else:
                    payment = existing_payment  # Usar el pago existente para la factura
                    payments_skipped += 1
                    self.stdout.write(self.style.WARNING(
                        f'    = Pago ya existe: {total_amount}€ - Factura {n_factura_str or "N/A"}'
                    ))
            elif total_amount > 0 and dry_run:
                payments_created += 1
                self.stdout.write(f'    [DRY] + Pago: {total_amount}€ - Factura {n_factura_str or "N/A"}')

            # Crear factura trimestral si hay número de factura
            if n_factura_str and total_amount > 0 and not dry_run and student:
                # Verificar si ya existe la factura
                existing_invoice = TaxInvoice.objects.filter(
                    invoice_number=n_factura_str
                ).first()

                if not existing_invoice:
                    # Detectar tasas
                    has_tasa_basica, has_tasa_a, has_traslado, renovaciones = self.detect_tasas(tasas_amount)

                    # Determinar trimestre y año
                    invoice_date = fecha_parsed.date() if fecha_parsed else datetime.now().date()
                    quarter = self.get_quarter_from_date(invoice_date)
                    year = invoice_date.year

                    # Crear factura
                    tax_invoice = TaxInvoice.objects.create(
                        student=student,
                        invoice_number=n_factura_str,
                        fecha=invoice_date,
                        quarter=quarter,
                        year=year,
                        curso=self.get_curso_code(curso),
                        has_tasa_basica=has_tasa_basica,
                        has_tasa_a=has_tasa_a,
                        has_traslado=has_traslado,
                        renovaciones_count=renovaciones,
                        base_imponible=base_amount,
                        iva_amount=iva_amount,
                        tasas_amount=tasas_amount,
                        total=total_amount,
                        client_name=f'{first_name} {last_name}',
                        client_dni=dni_clean,
                        client_street=str(direccion).strip() if direccion else '',
                        client_postal_code=str(cp).strip() if cp else '',
                        client_municipality=str(municipio).strip() if municipio else '',
                        client_province=str(provincia).strip() if provincia else 'VALENCIA',
                        notes=f'Importado de {excel_path.name}'
                    )

                    # Asociar pago a la factura
                    if payment:
                        tax_invoice.payments.add(payment)

                    invoices_created += 1
                    self.stdout.write(self.style.SUCCESS(
                        f'    + Factura: {n_factura_str} (Base: {base_amount}€, IVA: {iva_amount}€, Tasas: {tasas_amount}€)'
                    ))
                else:
                    invoices_skipped += 1
                    self.stdout.write(self.style.WARNING(
                        f'    = Factura ya existe: {n_factura_str}'
                    ))

            elif n_factura_str and total_amount > 0 and dry_run:
                invoices_created += 1
                self.stdout.write(f'    [DRY] + Factura: {n_factura_str}')

        wb.close()

        # Resumen
        self.stdout.write('')
        self.stdout.write('=' * 50)
        self.stdout.write(self.style.SUCCESS(f'Alumnos creados: {students_created}'))
        self.stdout.write(self.style.SUCCESS(f'Alumnos actualizados: {students_updated}'))
        self.stdout.write(self.style.SUCCESS(f'Pagos registrados: {payments_created}'))
        self.stdout.write(self.style.SUCCESS(f'Facturas creadas: {invoices_created}'))
        if payments_skipped:
            self.stdout.write(f'Pagos duplicados (ignorados): {payments_skipped}')
        if invoices_skipped:
            self.stdout.write(f'Facturas duplicadas (ignoradas): {invoices_skipped}')
        if rows_skipped:
            self.stdout.write(self.style.WARNING(f'Filas saltadas: {rows_skipped}'))
        self.stdout.write('=' * 50)

        if dry_run:
            self.stdout.write(self.style.WARNING(
                '\nMODO SIMULACION - Ejecutar sin --dry-run para guardar cambios'
            ))
